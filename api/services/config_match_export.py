"""Config match Excel export service.

Generates a formatted .xlsx report from config_match results.
Uses openpyxl directly — no pandas dependency.
"""

import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Colour constants ──────────────────────────────────────────────────────────

MERIDIAN_GREEN = "0D5639"
WHITE = "FFFFFF"
ERROR_FILL = "FFE8E8"
DEVIATION_FILL = "FFFDE0"
AMBIGUOUS_FILL = "E8F0FF"
HEADER_GREY = "F2F2F2"

# ── Style helpers ─────────────────────────────────────────────────────────────

_THIN_SIDE = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

_HEADER_FONT = Font(bold=True, color=WHITE, size=11)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor=MERIDIAN_GREEN)

_CLASSIFICATION_FILL: dict[str, PatternFill] = {
    "data_error": PatternFill(fill_type="solid", fgColor=ERROR_FILL),
    "config_deviation": PatternFill(fill_type="solid", fgColor=DEVIATION_FILL),
    "ambiguous": PatternFill(fill_type="solid", fgColor=AMBIGUOUS_FILL),
}


def _apply_header_style(cell: Any) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_data_style(cell: Any, classification: str) -> None:
    cell.fill = _CLASSIFICATION_FILL.get(
        classification, PatternFill(fill_type="solid", fgColor=WHITE)
    )
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _autofit_columns(ws: Any, min_width: int = 10, max_width: int = 60) -> None:
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, min_width),
                    min(len(str(cell.value)), max_width),
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def _write_detail_row(
    ws: Any,
    row_idx: int,
    match: dict,
    columns: list[str],
) -> None:
    clf = match.get("classification", "ambiguous")
    col_map = {
        "Module": match.get("module", ""),
        "Classification": match.get("classification", ""),
        "Check ID": match.get("check_id", ""),
        "Record Key": match.get("record_key", ""),
        "Field": match.get("field", ""),
        "Actual Value": match.get("actual_value", ""),
        "Rule Expectation": match.get("std_rule_expectation", ""),
        "Config Evidence": match.get("config_evidence", ""),
        "Recommended Action": match.get("recommended_action", ""),
        "SAP T-Code": match.get("sap_tcode", ""),
        "Fix Priority": match.get("fix_priority", ""),
    }
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=col_map.get(col_name, ""))
        _apply_data_style(cell, clf)


def _write_header_row(ws: Any, row_idx: int, columns: list[str]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
        _apply_header_style(cell)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(
    ws: Any,
    matches: list[dict],
    summary: dict,
    version_id: str,
) -> None:
    ws.title = "Summary"

    # Row 1 — title (merged A1:F1)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Meridian Config Match Report"
    title_cell.font = Font(bold=True, size=14, color=MERIDIAN_GREEN)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — version + date
    ws["A2"] = f"Version: {version_id[:8]}   Generated: {date.today().isoformat()}"
    ws["A2"].font = Font(size=10, color="666666")

    # Row 3 — blank

    # Row 4 — summary table header
    ws["A4"] = "Metric"
    ws["B4"] = "Count"
    for col in ("A4", "B4"):
        _apply_header_style(ws[col])

    total = (
        summary.get("data_errors", 0)
        + summary.get("config_deviations", 0)
        + summary.get("ambiguous", 0)
    )
    summary_rows = [
        ("Total Records Assessed", total),
        ("Data Errors", summary.get("data_errors", 0)),
        ("Config Deviations", summary.get("config_deviations", 0)),
        ("Ambiguous — Needs Review", summary.get("ambiguous", 0)),
    ]
    for i, (label, count) in enumerate(summary_rows, start=5):
        ws.cell(row=i, column=1, value=label).border = _THIN_BORDER
        ws.cell(row=i, column=2, value=count).border = _THIN_BORDER

    # Row 9 — blank
    # Row 10 — module breakdown header
    ws["A10"] = "Breakdown by Module"
    ws["A10"].font = Font(bold=True, size=11)

    module_cols = ["Module", "Data Errors", "Config Deviations", "Ambiguous", "Total"]
    _write_header_row(ws, 11, module_cols)

    # Compute per-module counts from matches list
    module_counts: dict[str, dict[str, int]] = {}
    for m in matches:
        mod = m.get("module", "unknown")
        clf = m.get("classification", "ambiguous")
        if mod not in module_counts:
            module_counts[mod] = {"data_error": 0, "config_deviation": 0, "ambiguous": 0}
        module_counts[mod][clf] = module_counts[mod].get(clf, 0) + 1

    row = 12
    for mod in summary.get("modules_with_deviations", []):
        counts = module_counts.get(mod, {})
        de = counts.get("data_error", 0)
        cd = counts.get("config_deviation", 0)
        am = counts.get("ambiguous", 0)
        vals = [mod, de, cd, am, de + cd + am]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = _THIN_BORDER
        row += 1

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


_DATA_ERROR_COLS = [
    "Module", "Check ID", "Record Key", "Field", "Actual Value",
    "Rule Expectation", "Recommended Action", "SAP T-Code", "Fix Priority",
]

_DEVIATION_COLS = [
    "Module", "Check ID", "Record Key", "Field", "Actual Value",
    "Config Evidence", "Recommended Action", "SAP T-Code", "Fix Priority",
]


def _build_classification_sheet(
    ws: Any,
    title: str,
    rows: list[dict],
    columns: list[str],
) -> None:
    ws.title = title
    _write_header_row(ws, 1, columns)
    for i, match in enumerate(rows, start=2):
        _write_detail_row(ws, i, match, columns)
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _build_module_sheet(ws: Any, module: str, rows: list[dict]) -> None:
    ws.title = module[:31]
    columns = [
        "Classification", "Check ID", "Record Key", "Field", "Actual Value",
        "Config Evidence", "Recommended Action", "SAP T-Code", "Fix Priority",
    ]
    _write_header_row(ws, 1, columns)
    sorted_rows = sorted(rows, key=lambda m: (m.get("fix_priority", 9), m.get("classification", "")))
    for i, match in enumerate(sorted_rows, start=2):
        _write_detail_row(ws, i, match, columns)
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_config_match_excel(
    matches: list[dict],
    summary: dict,
    version_id: str,
) -> bytes:
    """Generate a formatted Excel report from config match results.

    Args:
        matches: List of classification dicts from the config matching agent.
        summary: Summary dict with data_errors, config_deviations, ambiguous,
                 and modules_with_deviations counts.
        version_id: Analysis version UUID (first 8 chars shown in report).

    Returns:
        Excel file contents as bytes.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, matches, summary, version_id)

    # ── Sheet 2: Data Errors ──────────────────────────────────────────────────
    data_errors = sorted(
        [m for m in matches if m.get("classification") == "data_error"],
        key=lambda m: (m.get("fix_priority", 9), m.get("module", "")),
    )
    ws_errors = wb.create_sheet("Data Errors")
    _build_classification_sheet(ws_errors, "Data Errors", data_errors, _DATA_ERROR_COLS)

    # ── Sheet 3: Config Deviations ────────────────────────────────────────────
    deviations = sorted(
        [m for m in matches if m.get("classification") == "config_deviation"],
        key=lambda m: (m.get("module", ""), m.get("fix_priority", 9)),
    )
    ws_deviations = wb.create_sheet("Config Deviations")
    _build_classification_sheet(ws_deviations, "Config Deviations", deviations, _DEVIATION_COLS)

    # ── Sheet 4: Ambiguous ────────────────────────────────────────────────────
    ambiguous = sorted(
        [m for m in matches if m.get("classification") == "ambiguous"],
        key=lambda m: (m.get("module", ""), m.get("fix_priority", 9)),
    )
    ws_ambiguous = wb.create_sheet("Ambiguous — Review")
    _build_classification_sheet(ws_ambiguous, "Ambiguous — Review", ambiguous, _DEVIATION_COLS)

    # ── Sheets 5+: Per-module ─────────────────────────────────────────────────
    module_matches: dict[str, list[dict]] = {}
    for m in matches:
        mod = m.get("module", "unknown")
        module_matches.setdefault(mod, []).append(m)

    for mod in sorted(module_matches.keys()):
        ws_mod = wb.create_sheet(mod[:31])
        _build_module_sheet(ws_mod, mod, module_matches[mod])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
