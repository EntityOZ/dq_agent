"""Smoke tests for the config match Excel export service."""

import io

import openpyxl
import pytest

from api.services.config_match_export import generate_config_match_excel


def _make_match(
    classification: str,
    module: str,
    check_id: str = "BP-001",
    fix_priority: int = 2,
) -> dict:
    return {
        "module": module,
        "classification": classification,
        "check_id": check_id,
        "record_key": f"{module}-REC-001",
        "field": "SMTP_ADDR",
        "actual_value": "bad@example",
        "std_rule_expectation": "Valid email required",
        "config_evidence": "No email validation config found",
        "recommended_action": "Enable email validation rule",
        "sap_tcode": "BP",
        "fix_priority": fix_priority,
    }


def _make_summary(matches: list[dict]) -> dict:
    modules = list({m["module"] for m in matches})
    return {
        "data_errors": sum(1 for m in matches if m["classification"] == "data_error"),
        "config_deviations": sum(1 for m in matches if m["classification"] == "config_deviation"),
        "ambiguous": sum(1 for m in matches if m["classification"] == "ambiguous"),
        "modules_with_deviations": modules,
    }


@pytest.fixture
def ten_matches() -> list[dict]:
    return [
        _make_match("data_error", "business_partner", "BP-001"),
        _make_match("data_error", "business_partner", "BP-002"),
        _make_match("data_error", "fi_gl", "GL-001"),
        _make_match("data_error", "fi_gl", "GL-002"),
        _make_match("config_deviation", "business_partner", "BP-003"),
        _make_match("config_deviation", "material_master", "MM-001"),
        _make_match("config_deviation", "material_master", "MM-002"),
        _make_match("ambiguous", "business_partner", "BP-004"),
        _make_match("ambiguous", "fi_gl", "GL-003"),
        _make_match("ambiguous", "material_master", "MM-003"),
    ]


def test_excel_generates_without_error(ten_matches):
    summary = _make_summary(ten_matches)
    result = generate_config_match_excel(ten_matches, summary, "test-version-id")
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_excel_has_required_sheets(ten_matches):
    summary = _make_summary(ten_matches)
    result = generate_config_match_excel(ten_matches, summary, "test-version-id")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert "Summary" in wb.sheetnames
    assert "Data Errors" in wb.sheetnames
    assert "Config Deviations" in wb.sheetnames
    assert "Ambiguous — Review" in wb.sheetnames


def test_data_errors_sheet_row_count():
    matches = (
        [_make_match("data_error", "business_partner")] * 5
        + [_make_match("config_deviation", "fi_gl")] * 3
        + [_make_match("ambiguous", "material_master")] * 2
    )
    summary = _make_summary(matches)
    result = generate_config_match_excel(matches, summary, "test-version-id")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Data Errors"]
    # 1 header row + 5 data rows = 6
    assert ws.max_row == 6


def test_module_sheets_created_only_for_matched_modules():
    matches = [
        _make_match("data_error", "business_partner"),
        _make_match("config_deviation", "fi_gl"),
    ]
    summary = _make_summary(matches)
    result = generate_config_match_excel(matches, summary, "test-version-id")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert "business_partner" in wb.sheetnames
    assert "fi_gl" in wb.sheetnames
    assert "material_master" not in wb.sheetnames
